from collections import Counter
from re import X
from openpyxl import load_workbook 
wb = load_workbook("pairend1.xlsx") 
ws = wb.active 
source = wb["Sayfa1"]
print("sample verileri:")
listem=[]
for cell in source['A']:
 #print(" | "+str(cell.value)+" | ") 
 
 for i in range(0, len(['A'])):
   a= str(cell.value[0:3])
   listem.append(a[i])
print(listem)

b=listem.count('0/1')
print(b)



if(ws["C5"].value<0.10):
    print("erkek")
elif(ws["C5"].value>0.50):
    print("kadın")
else:
    print("unknown")